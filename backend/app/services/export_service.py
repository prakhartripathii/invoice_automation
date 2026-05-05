"""Render an Invoice (with merged OCR fields) to PDF or XLSX bytes.

The DB model only stores the structured columns (vendor_name, totals, items).
Extra UI fields (vendor_address, GST/IGST/CGST, terms, etc.) live inside
`invoice.champ_ocr_raw` as a JSON blob — we merge them in here so exports
match what the user sees on the detail screen.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict

from app.db.models.invoice import Invoice


def _merge_fields(inv: Invoice) -> Dict[str, Any]:
    raw: Dict[str, Any] = dict(inv.champ_ocr_raw or {})
    # Column values win over OCR raw — they're the reviewed/edited truth.
    overrides = {
        "vendor_name": inv.vendor_name,
        "invoice_number": inv.invoice_number,
        "invoice_date": inv.invoice_date,
        "due_date": inv.due_date,
        "currency": inv.currency,
        "subtotal": inv.subtotal,
        "tax_amount": inv.tax_amount,
        "total_amount": inv.total_amount,
        "purchase_order": inv.purchase_order,
    }
    for k, v in overrides.items():
        if v is not None:
            raw[k] = v
    return raw


def _fmt(v: Any) -> str:
    if v is None or v == "":
        return "—"
    return str(v)


# Field groups mirror the InvoiceDetail UI for parity
SUMMARY_GROUPS = [
    ("A. Basic Information", [
        ("invoice_number", "Invoice No."),
        ("invoice_date", "Date of Invoice"),
        ("purchase_order", "PO No."),
    ]),
    ("B. Vendor Details", [
        ("vendor_name", "Vendor Name"),
        ("vendor_address", "Address"),
        ("vendor_phone", "Phone Number"),
        ("vendor_email", "Email"),
        ("vendor_tax_id", "Tax ID / GSTIN"),
    ]),
    ("C. Financial Details", [
        ("total_quantity", "Total Quantity"),
        ("subtotal", "Subtotal"),
        ("gst", "GST"),
        ("igst", "IGST"),
        ("cgst", "CGST"),
        ("tax_amount", "Tax (combined)"),
        ("total_amount", "Total Amount"),
    ]),
    ("D. Additional Information", [
        ("terms_and_conditions", "Terms & Conditions"),
    ]),
]


def build_invoice_pdf(inv: Invoice) -> bytes:
    """Render a one-page (or short multi-page) PDF summary of the invoice."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    fields = _merge_fields(inv)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Invoice {fields.get('invoice_number') or inv.id}",
    )
    styles = getSampleStyleSheet()
    h_style = ParagraphStyle(
        "H", parent=styles["Heading2"], spaceAfter=6, textColor=colors.HexColor("#1f2937")
    )
    section_style = ParagraphStyle(
        "Sec",
        parent=styles["Heading4"],
        textColor=colors.HexColor("#2563eb"),
        spaceBefore=10,
        spaceAfter=4,
    )

    story: list = []
    story.append(Paragraph(
        f"Invoice {fields.get('invoice_number') or '—'}", h_style
    ))
    story.append(Paragraph(
        f"Status: <b>{inv.status.value}</b> &nbsp;·&nbsp; Uploaded: "
        f"{inv.created_at.strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 8))

    for title, group in SUMMARY_GROUPS:
        story.append(Paragraph(title, section_style))
        rows = [[label, _fmt(fields.get(key))] for key, label in group]
        t = Table(rows, colWidths=[55 * mm, 110 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#374151")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)

    if inv.items:
        story.append(Paragraph("Line Items", section_style))
        item_rows = [["#", "Description", "Qty", "Unit Price", "Amount"]]
        for it in sorted(inv.items, key=lambda x: x.line_number):
            item_rows.append([
                it.line_number,
                it.description,
                _fmt(it.quantity),
                _fmt(it.unit_price),
                _fmt(it.amount),
            ])
        t = Table(item_rows, colWidths=[12 * mm, 80 * mm, 20 * mm, 25 * mm, 28 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()


def build_invoice_xlsx(inv: Invoice) -> bytes:
    """Build a two-sheet workbook: Summary + Line Items."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    fields = _merge_fields(inv)
    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for title, group in SUMMARY_GROUPS:
        ws.append([title, ""])
        ws.cell(row=ws.max_row, column=1).font = bold
        for key, label in group:
            ws.append([label, _fmt(fields.get(key))])
    ws.append(["Status", inv.status.value])
    ws.append(["Uploaded", inv.created_at.strftime("%Y-%m-%d %H:%M")])
    if inv.posted_at:
        ws.append(["Posted at", inv.posted_at.strftime("%Y-%m-%d %H:%M")])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Line Items sheet ----
    items_ws = wb.create_sheet("Line Items")
    items_ws.append([
        "Line", "Description", "Quantity", "Unit Price", "Amount", "Tax Rate",
    ])
    for cell in items_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for it in sorted(inv.items or [], key=lambda x: x.line_number):
        items_ws.append([
            it.line_number,
            it.description,
            float(it.quantity) if it.quantity is not None else None,
            float(it.unit_price) if it.unit_price is not None else None,
            float(it.amount) if it.amount is not None else None,
            float(it.tax_rate) if it.tax_rate is not None else None,
        ])
    items_ws.column_dimensions["A"].width = 6
    items_ws.column_dimensions["B"].width = 50
    for col in ("C", "D", "E", "F"):
        items_ws.column_dimensions[col].width = 14

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
